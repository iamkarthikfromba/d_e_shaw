import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

# -------------- Functions --------------

def filter_dataframe(df, columns, filter_cond=None):
    new_df = df[columns].copy()

    if filter_cond:
        for col, (op, val) in filter_cond.items():
            if op == "==":
                new_df = new_df[new_df[col] == val]
            elif op == "!=":
                new_df = new_df[new_df[col] != val]

    new_df = new_df.fillna('-')
    return new_df

def style_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Report')
        workbook = writer.book
        worksheet = writer.sheets['Report']

        all_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_font = Font(name='Calibri', size=11, bold=True)
        regular_font = Font(name='Calibri', size=11)
        header_fill = PatternFill(start_color='#99C4ED', end_color='#99C4ED', fill_type="solid")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = all_border
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = regular_font

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        max_row = worksheet.max_row
        max_col = worksheet.max_column
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        worksheet.sheet_view.showGridLines = False

    output.seek(0)
    return output

# -------------- Streamlit App --------------

st.title("Cadence Report Generator 📊")

uploaded_file = st.file_uploader("Upload the Verified Master Roster file", type=["xlsx"])

if uploaded_file:
    df_master = pd.read_excel(uploaded_file)

    st.success("Master file loaded successfully!")

    report_options = {
    "HRIS Report to Finance": {
        "columns": [
            "Person ID",
            "Payroll ID",
            "First Name",
            "Last Name",
            "Login",
            "Primary Manager Name",
            "Primary Department",
            "Primary Sub-Department",
            "ProjectGroup1",
            "Job Title",
            "Level",
            "Work Status",
            "Source of Hire",
            "Latest Hire Date",
            "Location"
        ],
        "filter": None
    },
    "HRIS Report to BP": {
        "columns": [
            "Payroll ID",
            "First Name",
            "Last Name",
            "Login",
            "Gender",
            "Primary Manager Name",
            "Primary Manager Login",
            "Super Department",
            "Primary Department",
            "Primary Sub-Department",
            "ProjectGroup1",
            "Is Manager",
            "Job Title",
            "Level",
            "Work Status",
            "DESIS Stream",
            "Source of Hire",
            "Latest Hire Date",
            "Location"
        ],
        "filter": None
    },
    "HRIS Report to Tech BPs": {
        "columns": [
            "Payroll ID",
            "First Name",
            "Last Name",
            "Login",
            "Gender",
            "Primary Manager Name",
            "Primary Manager Login",
            "Super Department",
            "Primary Department",
            "Primary Sub-Department",
            "ProjectGroup1",
            "Is Manager",
            "Job Title",
            "Job Family",
            "Level",
            "Work Status",
            "DESIS Stream",
            "Source of Hire",
            "Latest Hire Date",
            "Location"
        ],
        "filter": {"Super Department": ["==", "Tech"]}
    },
    "HRIS Report to Non-Tech BPs": {
        "columns": [
            "Payroll ID",
            "First Name",
            "Last Name",
            "Login",
            "Gender",
            "Primary Manager Name",
            "Primary Manager Login",
            "Super Department",
            "Primary Department",
            "Primary Sub-Department",
            "ProjectGroup1",
            "Is Manager",
            "Job Title",
            "Level",
            "Work Status",
            "Source of Hire",
            "Latest Hire Date",
            "Location"
        ],
        "filter": {"Super Department": ["!=", "Tech"]}
    },
    "HRIS Report to TD": {
        "columns": [
            "Login",
            "First Name",
            "Last Name",
            "Full Name",
            "Primary Manager Name",
            "Primary Manager Login",
            "AD/AD+ Name",
            "AD/AD+ Login",
            "Primary Department",
            "Primary Sub-Department",
            "ProjectGroup1",
            "Is Manager",
            "Job Title",
            "Level",
            "Work Status",
            "DESIS Stream",
            "Location"
        ],
        "filter": None
    },
    "HRIS Report to EE": {
        "columns": [
            "Payroll ID",
            "Login",
            "First Name",
            "Last Name",
            "Gender",
            "Primary Department",
            "Job Title",
            "Level",
            "Work Status",
            "Latest Hire Date",
            "Location"
        ],
        "filter": None
    },
    "HRIS Report to DEI": {
        "columns": [
            "Login",
            "First Name",
            "Last Name",
            "Gender",
            "Primary Manager Login",
            "Super Department",
            "Primary Department",
            "Is Manager",
            "Job Title",
            "Level",
            "Work Status",
            "Latest Hire Date",
            "Location"
        ],
        "filter": None
    },
    "HRIS Report to ER": {
        "columns": [
            "Person ID",
            "Payroll ID",
            "First Name",
            "Last Name",
            "Full Name",
            "Login",
            "Gender",
            "Primary Manager Name",
            "Primary Manager Login",
            "Skip Manager Name",
            "OC-1",
            "OC-2",
            "Is Manager",
            "Super Department",
            "Primary Department",
            "Primary Sub-Department",
            "Job Title",
            "Level",
            "Work Status",
            "Latest Hire Date",
            "Location"
        ],
        "filter": None
    },
    "FTE Gender Headcount HRIS Report to Finance": {
        "columns": [
            "Payroll ID",
            "Login",
            "First Name",
            "Last Name",
            "Primary Department",
            "Job Title",
            "Gender",
            "Location",
            "Work Status"
        ],
        "filter": {"Work Status": ["==", "Regular FT"]}
    },
     "Contractor HRIS Report to Finance": {
        "columns": [
            "Payroll ID",
            "Login",
            "First Name",
            "Last Name",
            "Primary Department",
            "Job Title",
            "Gender",
            "Location",
            "Work Status"
        ],
        "filter": {"Work Status": ["==", "Contractor"]}
    }
}

    selected_report = st.selectbox("Select which report you want to generate:", list(report_options.keys()))

    if selected_report:
        config = report_options[selected_report]
        filtered_df = filter_dataframe(df_master, config["columns"], config["filter"])

        st.dataframe(filtered_df)

        styled_excel = style_excel(filtered_df)

        st.download_button(
            label="📥 Download Report as Excel",
            data=styled_excel,
            file_name=f"{selected_report.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.markdown("""
### Note:
- The naming convention for reporting purpose needs to be changed by the user.
- The date columns need to be converted to  short date in Excel.
- The report on Contractors needs to have an additional column with contractor company.
- The "FTE Gender Headcount HRIS Report to Finance" report and the "Contractor HRIS Report to Finance" have one more column "Work Status" which needs to be removed.

---

*Made by Balaraj B Gulabal, Contractor @ D E Shaw*
""")
